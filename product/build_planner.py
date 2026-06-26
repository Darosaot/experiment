#!/usr/bin/env python3
"""
Build the Content & Social Media Planner workbook — the second product,
designed to bundle with the Finance Toolkit for a higher-value sale.

    python3 build_planner.py

Output: Content-Social-Planner.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

from build_toolkit import (
    banner_font, txt, fill, set_widths, title_block, header_row, BORDER,
    INK, NAVY, TEAL, LIGHT, ACCENTBG, WHITE, MUTED, GOOD, WARN,
)
from openpyxl.styles import Font, Alignment

ROWS = 300


def build():
    wb = Workbook()

    # ---------------- Settings
    sett = wb.active
    sett.title = "Settings"
    title_block(sett, "⚙  Settings", "Edit these lists once — dropdowns everywhere else update.")
    set_widths(sett, {"A": 4, "B": 26, "C": 26, "D": 26, "E": 26})
    cols = {
        "B": ("Platforms", ["Instagram", "TikTok", "LinkedIn", "YouTube",
                             "Pinterest", "X / Twitter", "Blog", "Newsletter", "Facebook"]),
        "C": ("Content pillars", ["Educational", "Behind the scenes", "Promotion",
                                  "Testimonial", "Personal", "Trend / News", "Engagement"]),
        "D": ("Formats", ["Reel / Short", "Carousel", "Single image", "Story",
                          "Long video", "Article", "Text post", "Live"]),
        "E": ("Status", ["Idea", "Drafting", "Scheduled", "Published", "On hold"]),
    }
    ranges = {}
    for col, (title, items) in cols.items():
        c = sett[f"{col}4"]
        c.value = title
        c.font = banner_font(size=10)
        c.fill = fill(TEAL)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        for i, it in enumerate(items):
            cell = sett[f"{col}{5+i}"]
            cell.value = it
            cell.border = BORDER
            cell.fill = fill(LIGHT if i % 2 else WHITE)
        ranges[title] = f"Settings!${col}$5:${col}${4+len(items)}"

    # ---------------- Content Calendar
    cal = wb.create_sheet("Content Calendar")
    title_block(cal, "🗓  Content Calendar", "Your publishing plan. Status drives the Overview dashboard.")
    headers = ["Date", "Title / Hook", "Platform", "Pillar", "Format", "Status", "Link", "Notes"]
    set_widths(cal, {"A": 13, "B": 34, "C": 15, "D": 17, "E": 15, "F": 14, "G": 24, "H": 30})
    header_row(cal, 4, headers)
    for i in range(ROWS):
        r = 5 + i
        for col in range(1, 9):
            cell = cal.cell(row=r, column=col)
            cell.border = BORDER
            if i % 2:
                cell.fill = fill(LIGHT)
        cal.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    last = 4 + ROWS
    cal.freeze_panes = "A5"

    def dv(col, rng):
        v = DataValidation(type="list", formula1=f"={rng}", allow_blank=True)
        cal.add_data_validation(v)
        v.add(f"{col}5:{col}{last}")

    dv("C", ranges["Platforms"])
    dv("D", ranges["Content pillars"])
    dv("E", ranges["Formats"])
    dv("F", ranges["Status"])

    STATUS = f"'Content Calendar'!$F$5:$F${last}"
    PLAT = f"'Content Calendar'!$C$5:$C${last}"

    # ---------------- Idea Bank
    idea = wb.create_sheet("Idea Bank")
    title_block(idea, "💡  Idea Bank", "Park every idea here. Move the good ones into the Calendar.")
    set_widths(idea, {"A": 40, "B": 17, "C": 17, "D": 40})
    header_row(idea, 4, ["Idea / Hook", "Platform", "Pillar", "Notes"])
    for i in range(ROWS):
        r = 5 + i
        for col in range(1, 5):
            cell = idea.cell(row=r, column=col)
            cell.border = BORDER
            if i % 2:
                cell.fill = fill(LIGHT)
    v1 = DataValidation(type="list", formula1=f"={ranges['Platforms']}", allow_blank=True)
    idea.add_data_validation(v1)
    v1.add(f"B5:B{4+ROWS}")
    v2 = DataValidation(type="list", formula1=f"={ranges['Content pillars']}", allow_blank=True)
    idea.add_data_validation(v2)
    v2.add(f"C5:C{4+ROWS}")
    idea.freeze_panes = "A5"

    # ---------------- Overview dashboard
    ov = wb.create_sheet("Overview")
    title_block(ov, "📊  Overview", "Counts update as you fill the Content Calendar.")
    set_widths(ov, {"A": 4, "B": 24, "C": 14, "D": 6, "E": 22, "F": 12})

    # status counts
    ov["B4"] = "By status"
    ov["B4"].font = banner_font(size=11, color=NAVY)
    ov["B4"].fill = fill(ACCENTBG)
    statuses = cols["E"][1]
    for i, s in enumerate(statuses):
        r = 5 + i
        ov.cell(row=r, column=2, value=s).border = BORDER
        cnt = ov.cell(row=r, column=3, value=f'=COUNTIF({STATUS},$B{r})')
        cnt.border = BORDER
        cnt.alignment = Alignment(horizontal="center")
        cnt.font = txt(bold=True, color=NAVY)

    # platform counts (drives chart)
    ov["E4"] = "By platform"
    ov["E4"].font = banner_font(size=11, color=NAVY)
    ov["E4"].fill = fill(ACCENTBG)
    plats = cols["B"][1]
    for i, p in enumerate(plats):
        r = 5 + i
        ov.cell(row=r, column=5, value=p).border = BORDER
        cnt = ov.cell(row=r, column=6, value=f'=COUNTIF({PLAT},$E{r})')
        cnt.border = BORDER
        cnt.alignment = Alignment(horizontal="center")
        cnt.font = txt(bold=True, color=NAVY)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Posts by platform"
    chart.height = 9
    chart.width = 14
    data = Reference(ov, min_col=6, min_row=4, max_row=4 + len(plats))
    cats = Reference(ov, min_col=5, min_row=5, max_row=4 + len(plats))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ov.add_chart(chart, "B14")

    # total posts planned
    ov["B12"] = "Total posts planned"
    ov["B12"].font = txt(bold=True)
    tot = ov["C12"]
    tot.value = f'=COUNTA({PLAT})'
    tot.font = Font(name="Calibri", size=14, bold=True, color=TEAL)
    tot.alignment = Alignment(horizontal="center")

    # ---------------- Start Here
    home = wb.create_sheet("Start Here")
    wb.move_sheet("Start Here", -(len(wb.sheetnames) - 1))
    title_block(home, "Content & Social Media Planner",
                "Plan a month of content in minutes — for any platform.", last_col="H")
    home.sheet_view.showGridLines = False
    set_widths(home, {"A": 4, "B": 64})
    lines = [
        ("", ""),
        ("HOW TO USE", "h"),
        ("1.  Open  Settings  and tweak your platforms, pillars and formats.", ""),
        ("2.  Dump every idea into the  Idea Bank  (no pressure, no order).", ""),
        ("3.  Schedule the best ones on the  Content Calendar  with a date & status.", ""),
        ("4.  Watch the  Overview  tab count your plan by status and platform.", ""),
        ("", ""),
        ("WORKS IN", "h"),
        ("Microsoft Excel · Google Sheets · Apple Numbers · LibreOffice Calc", ""),
        ("", ""),
        ("TIP", "h"),
        ("Batch one content type at a time (e.g. all reels) — it's far faster.", ""),
        ("", ""),
        ("Pairs perfectly with the Freelancer Finance & Business Toolkit.", "f"),
    ]
    r = 4
    for text, kind in lines:
        cell = home.cell(row=r, column=2, value=text)
        if kind == "h":
            cell.font = banner_font(size=12, color=NAVY)
            cell.fill = fill(ACCENTBG)
        elif kind == "f":
            cell.font = txt(italic=True, color=TEAL, bold=True)
        else:
            cell.font = txt(size=11)
        home.row_dimensions[r].height = 18
        r += 1

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "Content-Social-Planner.xlsx")
    wb.save(out)
    print("Saved:", out)
    return out


if __name__ == "__main__":
    build()
