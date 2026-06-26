#!/usr/bin/env python3
"""
Build the Freelancer Finance & Business Toolkit workbook.

Generates a polished, multi-tab .xlsx that works in both Microsoft Excel
and Google Sheets. Re-run this script any time to regenerate the product
file from scratch:

    python3 build_toolkit.py

Output: Freelancer-Finance-Toolkit.xlsx (next to this script)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import os

# ---------------------------------------------------------------- palette
INK      = "1F2933"   # near-black text
NAVY     = "1B3A5B"   # headers / banner
TEAL     = "2C7A7B"   # accents
LIGHT    = "EDF2F7"   # zebra / panels
ACCENTBG = "E6FFFA"   # highlight panels
WHITE    = "FFFFFF"
MUTED    = "718096"
GOOD     = "276749"
WARN     = "9C4221"

THIN = Side(style="thin", color="CBD5E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_ROWS = 250  # generous entry space


def banner_font(size=11, color=WHITE, bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def txt(size=11, color=INK, bold=False, italic=False):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, subtitle, last_col="H"):
    """Top banner used on every sheet."""
    ws.merge_cells(f"A1:{last_col}1")
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A1"]
    c.value = title
    c.font = banner_font(size=18)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s = ws["A2"]
    s.value = subtitle
    s.font = txt(size=10, color=WHITE, italic=True)
    s.fill = fill(NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 18
    ws.sheet_view.showGridLines = False


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = banner_font(size=10)
        cell.fill = fill(TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


# ===================================================================
def build():
    wb = Workbook()

    # ---- currency style (driven by Settings, but use a generic accounting fmt)
    money_fmt = '#,##0.00'

    # =============================================== SETTINGS (build first)
    sett = wb.active
    sett.title = "Settings"
    title_block(sett, "⚙  Settings", "Set these once. Everything else updates automatically.")
    set_widths(sett, {"A": 4, "B": 34, "C": 22, "D": 50})

    rows = [
        ("Your name / business", "Your Name", "Shown for your own reference."),
        ("Currency symbol", "€", "Type any symbol: €, $, £, etc."),
        ("Annual income goal", 30000, "Used by the Rate Calculator."),
        ("Billable hours per year", 1200, "Realistic billable hours (not total worked)."),
        ("Tax set-aside rate (%)", 25, "Rough % to reserve for income tax + social/self-employment."),
        ("Business expenses / year (est.)", 4000, "Used by the Rate Calculator."),
    ]
    sett["B4"] = "SETTING"
    sett["C4"] = "VALUE"
    sett["D4"] = "NOTES"
    for col in ("B", "C", "D"):
        cell = sett[f"{col}4"]
        cell.font = banner_font(size=10)
        cell.fill = fill(TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    r = 5
    for label, val, note in rows:
        sett.cell(row=r, column=2, value=label).font = txt(bold=True)
        vcell = sett.cell(row=r, column=3, value=val)
        vcell.font = txt(color=NAVY, bold=True)
        vcell.alignment = Alignment(horizontal="center")
        vcell.fill = fill(ACCENTBG)
        sett.cell(row=r, column=4, value=note).font = txt(color=MUTED, italic=True)
        for col in (2, 3, 4):
            sett.cell(row=r, column=col).border = BORDER
        r += 1

    # named cells for reuse
    GOAL = "Settings!$C$7"
    HOURS = "Settings!$C$8"
    TAXRATE = "Settings!$C$9"   # percent number, e.g. 25
    EXP_EST = "Settings!$C$10"

    # category list for validation / reference
    sett["B13"] = "Expense categories (edit freely):"
    sett["B13"].font = txt(bold=True)
    categories = [
        "Software & subscriptions", "Hardware & equipment", "Office & supplies",
        "Travel & transport", "Marketing & advertising", "Professional services",
        "Education & training", "Bank & payment fees", "Insurance",
        "Rent & utilities", "Meals & entertainment", "Other",
    ]
    for i, cat in enumerate(categories):
        cell = sett.cell(row=14 + i, column=2, value=cat)
        cell.border = BORDER
        cell.fill = fill(LIGHT if i % 2 else WHITE)
    cat_range = f"Settings!$B$14:$B${14 + len(categories) - 1}"

    status_list = ["Paid", "Unpaid", "Overdue", "Draft"]
    sett["D13"] = "Invoice statuses:"
    sett["D13"].font = txt(bold=True)
    for i, st in enumerate(status_list):
        cell = sett.cell(row=14 + i, column=4, value=st)
        cell.border = BORDER
        cell.fill = fill(LIGHT if i % 2 else WHITE)
    status_range = f"Settings!$D$14:$D${14 + len(status_list) - 1}"

    proj_status = ["Lead", "Proposal sent", "Active", "On hold", "Completed", "Cancelled"]
    sett["D20"] = "Project statuses:"
    sett["D20"].font = txt(bold=True)
    for i, st in enumerate(proj_status):
        cell = sett.cell(row=21 + i, column=4, value=st)
        cell.border = BORDER
        cell.fill = fill(LIGHT if i % 2 else WHITE)
    proj_status_range = f"Settings!$D$21:$D${21 + len(proj_status) - 1}"

    # =============================================== INCOME / INVOICES
    inc = wb.create_sheet("Income")
    title_block(inc, "💶  Income & Invoices", "Log every invoice. 'Paid' rows feed your Dashboard automatically.")
    inc_headers = ["Date", "Client", "Invoice #", "Description", "Amount", "Status", "Date paid"]
    set_widths(inc, {"A": 13, "B": 22, "C": 12, "D": 34, "E": 14, "F": 13, "G": 13})
    header_row(inc, 4, inc_headers)
    for i in range(DATA_ROWS):
        row = 5 + i
        for col in range(1, 8):
            cell = inc.cell(row=row, column=col)
            cell.border = BORDER
            if i % 2:
                cell.fill = fill(LIGHT)
        inc.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        inc.cell(row=row, column=5).number_format = money_fmt
        inc.cell(row=row, column=7).number_format = "yyyy-mm-dd"
    last_inc = 4 + DATA_ROWS
    # status dropdown
    dv = DataValidation(type="list", formula1=f"={status_range}", allow_blank=True)
    inc.add_data_validation(dv)
    dv.add(f"F5:F{last_inc}")

    INC_AMT = f"Income!$E$5:$E${last_inc}"
    INC_STATUS = f"Income!$F$5:$F${last_inc}"
    INC_DATE = f"Income!$A$5:$A${last_inc}"

    # =============================================== EXPENSES
    exp = wb.create_sheet("Expenses")
    title_block(exp, "🧾  Expenses", "Track costs by category. Deductible totals feed your tax estimate.")
    exp_headers = ["Date", "Vendor", "Category", "Description", "Amount", "Deductible?"]
    set_widths(exp, {"A": 13, "B": 22, "C": 24, "D": 30, "E": 14, "F": 13})
    header_row(exp, 4, exp_headers)
    for i in range(DATA_ROWS):
        row = 5 + i
        for col in range(1, 7):
            cell = exp.cell(row=row, column=col)
            cell.border = BORDER
            if i % 2:
                cell.fill = fill(LIGHT)
        exp.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        exp.cell(row=row, column=5).number_format = money_fmt
    last_exp = 4 + DATA_ROWS
    dv_cat = DataValidation(type="list", formula1=f"={cat_range}", allow_blank=True)
    exp.add_data_validation(dv_cat)
    dv_cat.add(f"C5:C{last_exp}")
    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    exp.add_data_validation(dv_yn)
    dv_yn.add(f"F5:F{last_exp}")

    EXP_AMT = f"Expenses!$E$5:$E${last_exp}"
    EXP_CAT = f"Expenses!$C$5:$C${last_exp}"
    EXP_DED = f"Expenses!$F$5:$F${last_exp}"
    EXP_DATE = f"Expenses!$A$5:$A${last_exp}"

    # =============================================== PROJECTS
    proj = wb.create_sheet("Projects")
    title_block(proj, "📁  Project Tracker", "Pipeline + delivery at a glance. Value = Quoted rate × Est. hours.")
    proj_headers = ["Project", "Client", "Status", "Est. hours", "Hourly rate",
                    "Quoted value", "Actual hours", "Effective rate"]
    set_widths(proj, {"A": 26, "B": 20, "C": 16, "D": 11, "E": 12, "F": 14, "G": 12, "H": 14})
    header_row(proj, 4, proj_headers)
    for i in range(DATA_ROWS):
        row = 5 + i
        for col in range(1, 9):
            cell = proj.cell(row=row, column=col)
            cell.border = BORDER
            if i % 2:
                cell.fill = fill(LIGHT)
        proj.cell(row=row, column=5).number_format = money_fmt
        # quoted value = est hours * rate
        proj.cell(row=row, column=6).value = f"=IF(AND($D{row}<>\"\",$E{row}<>\"\"),$D{row}*$E{row},\"\")"
        proj.cell(row=row, column=6).number_format = money_fmt
        # effective rate = quoted value / actual hours
        proj.cell(row=row, column=8).value = f"=IF(AND($F{row}<>\"\",$G{row}>0),$F{row}/$G{row},\"\")"
        proj.cell(row=row, column=8).number_format = money_fmt
    last_proj = 4 + DATA_ROWS
    dv_ps = DataValidation(type="list", formula1=f"={proj_status_range}", allow_blank=True)
    proj.add_data_validation(dv_ps)
    dv_ps.add(f"C5:C{last_proj}")

    # =============================================== RATE CALCULATOR
    rate = wb.create_sheet("Rate Calculator")
    title_block(rate, "🧮  Rate Calculator", "What must you charge per hour to hit your goal? Edit Settings to change inputs.", last_col="E")
    set_widths(rate, {"A": 4, "B": 38, "C": 18, "D": 4, "E": 4})

    def kv(row, label, formula, fmt=money_fmt, note=None, highlight=False):
        rate.cell(row=row, column=2, value=label).font = txt(bold=True)
        v = rate.cell(row=row, column=3, value=formula)
        v.number_format = fmt
        v.alignment = Alignment(horizontal="center")
        v.font = txt(color=NAVY, bold=True)
        v.border = BORDER
        rate.cell(row=row, column=2).border = BORDER
        if highlight:
            v.fill = fill(ACCENTBG)
            rate.cell(row=row, column=2).fill = fill(ACCENTBG)
        if note:
            n = rate.cell(row=row, column=4, value=note)
            n.font = txt(color=MUTED, italic=True)

    rate["B4"] = "Your inputs come from the Settings tab."
    rate["B4"].font = txt(color=MUTED, italic=True)
    kv(5, "Annual income goal", f"={GOAL}")
    kv(6, "Estimated annual expenses", f"={EXP_EST}")
    kv(7, "Tax set-aside rate", f"={TAXRATE}/100", fmt="0%")
    kv(8, "Billable hours per year", f"={HOURS}", fmt="#,##0")
    rate["B9"] = ""
    kv(10, "Revenue needed (pre-tax)",
       f"=({GOAL}+{EXP_EST})/(1-{TAXRATE}/100)",
       note="Gross billings to net your goal after expenses & tax.")
    kv(11, "→ Required hourly rate",
       f"=IF({HOURS}>0,(({GOAL}+{EXP_EST})/(1-{TAXRATE}/100))/{HOURS},0)",
       note="Charge at least this. Round up.", highlight=True)
    kv(12, "→ Required day rate (8h)",
       f"=IF({HOURS}>0,8*((({GOAL}+{EXP_EST})/(1-{TAXRATE}/100))/{HOURS}),0)",
       note="A clean number for proposals.", highlight=True)
    rate.cell(row=11, column=2).font = txt(bold=True, size=12)
    rate.cell(row=12, column=2).font = txt(bold=True, size=12)

    # =============================================== DASHBOARD (build last, ref others)
    dash = wb.create_sheet("Dashboard")
    title_block(dash, "📊  Dashboard", "Live snapshot of your business. Numbers update as you log income & expenses.")
    set_widths(dash, {"A": 4, "B": 30, "C": 18, "D": 6, "E": 26, "F": 16, "G": 6})

    # KPI cards
    def card(anchor_row, anchor_col, label, formula, fmt=money_fmt, color=NAVY):
        lc = dash.cell(row=anchor_row, column=anchor_col, value=label)
        lc.font = txt(size=10, color=MUTED, bold=True)
        lc.fill = fill(LIGHT)
        lc.alignment = Alignment(horizontal="left", indent=1)
        lc.border = BORDER
        vc = dash.cell(row=anchor_row + 1, column=anchor_col, value=formula)
        vc.font = Font(name="Calibri", size=16, bold=True, color=color)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="left", indent=1)
        vc.fill = fill(WHITE)
        vc.border = BORDER
        # span label+value across two columns visually
        dash.merge_cells(start_row=anchor_row, start_column=anchor_col,
                         end_row=anchor_row, end_column=anchor_col + 1)
        dash.merge_cells(start_row=anchor_row + 1, start_column=anchor_col,
                         end_row=anchor_row + 1, end_column=anchor_col + 1)
        dash.cell(row=anchor_row, column=anchor_col + 1).border = BORDER
        dash.cell(row=anchor_row + 1, column=anchor_col + 1).border = BORDER

    paid_income = f'SUMIF({INC_STATUS},"Paid",{INC_AMT})'
    outstanding = f'SUMIF({INC_STATUS},"Unpaid",{INC_AMT})+SUMIF({INC_STATUS},"Overdue",{INC_AMT})'
    total_exp = f"SUM({EXP_AMT})"
    ded_exp = f'SUMIF({EXP_DED},"Yes",{EXP_AMT})'
    net_profit = f"({paid_income}-{total_exp})"
    taxable = f"({paid_income}-{ded_exp})"
    tax_due = f"MAX(0,{taxable})*{TAXRATE}/100"
    take_home = f"({net_profit}-({tax_due}))"

    card(4, 2, "Income received (Paid)", f"={paid_income}", color=GOOD)
    card(4, 5, "Outstanding invoices", f"={outstanding}", color=WARN)
    card(7, 2, "Total expenses", f"={total_exp}", color=WARN)
    card(7, 5, "Net profit", f"={net_profit}", color=NAVY)
    card(10, 2, "Estimated tax to set aside", f"={tax_due}", color=WARN)
    card(10, 5, "Estimated take-home", f"={take_home}", color=GOOD)

    # progress vs goal
    dash["B14"] = "Progress to annual income goal"
    dash["B14"].font = txt(bold=True)
    dash["B15"] = f"=IF({GOAL}>0,{paid_income}/{GOAL},0)"
    dash["B15"].number_format = "0.0%"
    dash["B15"].font = Font(name="Calibri", size=20, bold=True, color=TEAL)
    dash["C14"] = "Goal"
    dash["C14"].font = txt(size=10, color=MUTED)
    dash["C15"] = f"={GOAL}"
    dash["C15"].number_format = money_fmt
    dash["C15"].font = txt(bold=True)

    # ---- monthly breakdown helper table (drives chart)
    dash["E18"] = "Month"
    dash["F18"] = "Income"
    dash["G18"] = "Expenses"
    for c in ("E18", "F18", "G18"):
        dash[c].font = banner_font(size=10)
        dash[c].fill = fill(TEAL)
        dash[c].alignment = Alignment(horizontal="center")
        dash[c].border = BORDER
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, m in enumerate(months):
        r = 19 + i
        dash.cell(row=r, column=5, value=m).border = BORDER
        # income by month (paid only)
        inc_f = (f'=SUMPRODUCT(({INC_STATUS}="Paid")*'
                 f'(MONTH({INC_DATE})={i+1})*(ISNUMBER({INC_DATE}))*{INC_AMT})')
        ec = dash.cell(row=r, column=6, value=inc_f)
        ec.number_format = money_fmt
        ec.border = BORDER
        exp_f = (f'=SUMPRODUCT((MONTH({EXP_DATE})={i+1})*'
                 f'(ISNUMBER({EXP_DATE}))*{EXP_AMT})')
        xc = dash.cell(row=r, column=7, value=exp_f)
        xc.number_format = money_fmt
        xc.border = BORDER

    # bar chart income vs expenses by month
    chart = BarChart()
    chart.type = "col"
    chart.title = "Income vs Expenses by month"
    chart.height = 7.5
    chart.width = 16
    data = Reference(dash, min_col=6, max_col=7, min_row=18, max_row=30)
    cats = Reference(dash, min_col=5, min_row=19, max_row=30)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.numFmt = '#,##0'
    dash.add_chart(chart, "B18")

    # expense-by-category table + pie (placed lower)
    start = 33
    dash.cell(row=start, column=5, value="Category").font = banner_font(size=10)
    dash.cell(row=start, column=6, value="Spent").font = banner_font(size=10)
    dash.cell(row=start, column=5).fill = fill(TEAL)
    dash.cell(row=start, column=6).fill = fill(TEAL)
    dash.cell(row=start, column=5).border = BORDER
    dash.cell(row=start, column=6).border = BORDER
    for i, cat in enumerate(categories):
        r = start + 1 + i
        dash.cell(row=r, column=5, value=cat).border = BORDER
        f = f'=SUMIF({EXP_CAT},$E{r},{EXP_AMT})'
        cc = dash.cell(row=r, column=6, value=f)
        cc.number_format = money_fmt
        cc.border = BORDER
    pie = PieChart()
    pie.title = "Expenses by category"
    pie.height = 8
    pie.width = 12
    pdata = Reference(dash, min_col=6, min_row=start, max_row=start + len(categories))
    plabels = Reference(dash, min_col=5, min_row=start + 1, max_row=start + len(categories))
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(plabels)
    dash.add_chart(pie, "B33")

    # =============================================== START HERE (front sheet)
    home = wb.create_sheet("Start Here")
    wb.move_sheet("Start Here", -(len(wb.sheetnames) - 1))  # move to first
    title_block(home, "Freelancer Finance & Business Toolkit",
                "A complete, automatic money dashboard for freelancers & solo businesses.", last_col="H")
    home.sheet_view.showGridLines = False
    set_widths(home, {"A": 4, "B": 60, "C": 30})
    lines = [
        ("", ""),
        ("HOW TO USE THIS TOOLKIT", "h"),
        ("1.  Open the  Settings  tab and fill in your goal, currency and tax rate.", ""),
        ("2.  Log invoices on the  Income  tab. Mark each one Paid / Unpaid.", ""),
        ("3.  Log costs on the  Expenses  tab and pick a category.", ""),
        ("4.  Watch the  Dashboard  update itself — profit, tax to set aside, take-home.", ""),
        ("5.  Use the  Rate Calculator  to know exactly what to charge.", ""),
        ("6.  Track your pipeline on the  Projects  tab.", ""),
        ("", ""),
        ("WHAT'S INSIDE", "h"),
        ("• Auto Dashboard with charts (income vs expenses, expenses by category)", ""),
        ("• Invoice tracker with Paid / Unpaid / Overdue status", ""),
        ("• Expense tracker with deductible flag for easy tax time", ""),
        ("• Tax set-aside estimator so you're never caught short", ""),
        ("• Hourly & day-rate calculator based on your real goals", ""),
        ("• Project / pipeline tracker", ""),
        ("", ""),
        ("WORKS IN", "h"),
        ("• Microsoft Excel (desktop & 365)", ""),
        ("• Google Sheets — upload to Drive, then File ▸ 'Save as Google Sheets'", ""),
        ("• Apple Numbers & LibreOffice Calc", ""),
        ("", ""),
        ("TIPS", "h"),
        ("• Dates: type them as 2026-01-31 so the monthly charts work.", ""),
        ("• Add rows by copying a formatted row down — formulas extend automatically.", ""),
        ("• Edit the category & status lists on the Settings tab to fit your work.", ""),
        ("", ""),
        ("Thank you for your purchase! Questions? Reply to your order receipt.", "f"),
    ]
    r = 4
    for text, kind in lines:
        cell = home.cell(row=r, column=2, value=text)
        if kind == "h":
            cell.font = banner_font(size=12, color=NAVY)
            home.cell(row=r, column=2).fill = fill(ACCENTBG)
            home.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        elif kind == "f":
            cell.font = txt(italic=True, color=TEAL, bold=True)
        else:
            cell.font = txt(size=11)
        home.row_dimensions[r].height = 18
        r += 1

    # cosmetic: freeze header rows on data sheets
    for name in ("Income", "Expenses", "Projects"):
        wb[name].freeze_panes = "A5"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "Freelancer-Finance-Toolkit.xlsx")
    wb.save(out)
    print("Saved:", out)
    return out


if __name__ == "__main__":
    build()
